import io
from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# Function: xlsx_to_text
# Purpose: Extract all text from an XLSX file-like object and return as a string
# Inputs:
#   - file: a FileStorage or file-like object representing the uploaded .xlsx
# Outputs:
#   - A single string containing concatenated text from all sheets and rows
# ----------------------------------------------------------------------------
def xlsx_to_text(file):
    # Read file bytes
    data = file.read()
    # Reset file pointer
    try:
        file.seek(0)
    except Exception:
        pass
    # Load workbook in read-only mode
    wb = load_workbook(io.BytesIO(data), read_only=True)
    texts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            # Join non-empty cell values
            row_vals = [str(cell) for cell in row if cell is not None]
            if row_vals:
                texts.append(' '.join(row_vals))
    return "\n\n".join(texts) 